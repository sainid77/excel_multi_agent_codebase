from __future__ import annotations

import zipfile
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.workbook.defined_name import DefinedName

from insurance_excel_agents.models import FormulaRecord, SheetInfo


class WorkbookParser:
    """Extract workbook metadata, sheet inventory, formulas, and names."""

    def inspect_container(self, path: Path) -> dict[str, Any]:
        with zipfile.ZipFile(path) as zf:
            names = set(zf.namelist())
        return {
            "has_embedded_vba": any(name.endswith("vbaProject.bin") for name in names),
            "has_connections": any("connections" in name.lower() for name in names),
            "parts": sorted(list(names))[:50],
        }

    def parse_workbook(self, path: Path) -> dict[str, Any]:
        wb = load_workbook(path, data_only=False, keep_vba=True)
        sheet_infos: list[SheetInfo] = []
        formulas: list[FormulaRecord] = []
        hidden_sheets: list[str] = []

        for ws in wb.worksheets:
            formula_count = 0
            if ws.sheet_state != "visible":
                hidden_sheets.append(ws.title)
            for row in ws.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str) and cell.value.startswith("="):
                        formula_count += 1
                        formulas.append(
                            FormulaRecord(
                                workbook=path.name,
                                sheet=ws.title,
                                cell=cell.coordinate,
                                formula=cell.value,
                            )
                        )
            sheet_infos.append(
                SheetInfo(
                    name=ws.title,
                    max_row=ws.max_row,
                    max_column=ws.max_column,
                    hidden=ws.sheet_state != "visible",
                    formulas=formula_count,
                )
            )

        named_ranges = self._extract_named_ranges(wb.defined_names)
        external_links = self._extract_external_links(wb)

        return {
            "workbook": path.name,
            "sheet_infos": [item.model_dump() for item in sheet_infos],
            "formulas": [item.model_dump() for item in formulas],
            "named_ranges": named_ranges,
            "hidden_sheets": hidden_sheets,
            "external_links": external_links,
            "container": self.inspect_container(path),
        }

    @staticmethod
    def _extract_named_ranges(defined_names: Any) -> list[dict[str, Any]]:
        results: list[dict[str, Any]] = []
        items = defined_names.values() if hasattr(defined_names, "values") else defined_names
        for dn in items:
            results.append(
                {
                    "name": dn.name,
                    "attr_text": dn.attr_text,
                    "is_external": getattr(dn, "is_external", False),
                }
            )
        return results

    @staticmethod
    def _extract_external_links(wb: Any) -> list[str]:
        links = []
        for link in getattr(wb, "_external_links", []) or []:
            target = getattr(link, "file_link", None)
            if target is not None:
                links.append(str(target))
            else:
                links.append(str(link))
        return links
